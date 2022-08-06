import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

print("Printing file:",inv_file.__dict__)

products_per_suplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    #calculation of number of products per supplier
    if supplier_name in products_per_suplier:
        current_num_products= products_per_suplier.get(supplier_name)
        products_per_suplier[supplier_name] = current_num_products + 1
    else:
        print("adding a new supplier")
        products_per_suplier[supplier_name] = 1
    #calculation total value of supplier

    if supplier_name in total_value_per_supplier:
        curent_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = curent_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    #logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[product_num]= inventory
    # add value on the excell
    inventory_price.value = inventory * price

print(products_per_suplier)
print(total_value_per_supplier)
print(products_under_10_inv)
inv_file.save("inventory_with_total_value.xlsx")



